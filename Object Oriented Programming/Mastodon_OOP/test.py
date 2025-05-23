import unittest
import MastodonOOP as MastodonOOPsolution
from mastodon import Mastodon
from bs4 import BeautifulSoup
import openpyxl
import os
import datetime


class Mastodon_api(unittest.TestCase):

    def test_API(self):
        self.assertIsInstance(MastodonOOPsolution.mastodon, Mastodon,
                              "Your API-Initiation does not correctly work, check again if you are missing anything!")


class Mastodon_test(unittest.TestCase):

    def setUp(self):
        self.toot_true = MastodonOOPsolution.Toot(
            account=[{"id": 123, "username": "Marco"}],
            toot_id=True,
            content='<p>Hello from Python, dog</p>',
            user_id=True,
            hashtags=[{'name': 'dog', 'url': True, 'history': ''}],
            bookmark=True,
            no_replies=True,
            url=True,
            count_replies=True,
            pubdate='2022-07-22 09:37:34+00:00',
            mentions=True,
            media=[{"id": 123, "type": "image"}, {"id": 1234, "type": "video"}, {"id": 12345, "type": "gifv"},
                   {"id": 123456, "type": "audio"}, {"id": 1234567, "type": "unknown"}],
            language='en',
            poll=True
        )
        self.toot_false = MastodonOOPsolution.Toot(
            account='',
            toot_id='',
            content='sun',
            user_id='',
            hashtags='',
            bookmark='',
            no_replies='',
            url='',
            count_replies='',
            pubdate='2024-07-22 09:37:34+00:00',
            mentions=False,
            media='',
            language='',
            poll=False
        )

        self.toot_true.pubdate = datetime.datetime.strptime(self.toot_true.pubdate, "%Y-%m-%d %H:%M:%S%z")
        self.toot_false.pubdate = datetime.datetime.strptime(self.toot_false.pubdate, "%Y-%m-%d %H:%M:%S%z")

        self.clock = '2023-07-22 09:37:34+00:00'

    def tearDown(self):
        pass

    def test_Toot(self):
        assert all(hasattr(self.toot_true, attr) for attr in ["content", "account", "toot_id", "user_id", "hashtags", "bookmark", "no_replies", "url", "count_replies",
                                                              "pubdate", "mentions", "media", "language", "poll"]), "Your Toot-Class does not correctly work, check again if you are missing one or more attributes!"

    def test_load(self):
        toots_dict = []
        hashtag = "Moin"
        mastodon = Mastodon(
            client_id=MastodonOOPsolution.mastodon.client_id,
            client_secret=MastodonOOPsolution.mastodon.client_secret,
            access_token=MastodonOOPsolution.mastodon.access_token,
            api_base_url=MastodonOOPsolution.mastodon.api_base_url
        )

        # Load all toots with a specific hashtag into a dictionary, limit to 10 toots
        toots = mastodon.timeline_hashtag(hashtag, limit=10)
        result = MastodonOOPsolution.load(hashtag)

        # Process the retrieved toots
        for toot in toots:
            content_html = toot['content']
            soup = BeautifulSoup(content_html, 'html.parser')
            content_text = soup.get_text()
            toot = MastodonOOPsolution.Toot(
                account=toot['account'],
                toot_id=toot['id'],
                content=content_text,
                user_id=toot['account']['id'],
                hashtags=toot['tags'],
                bookmark=toot['bookmarked'],
                no_replies=toot['reblogs_count'],
                url=toot['url'],
                count_replies=toot['replies_count'],
                pubdate=toot['created_at'],
                mentions=toot['mentions'],
                media=toot['media_attachments'],
                language=toot['language'],
                poll=toot['poll']
            )
            toots_dict.append(toot)

        true_bool = True

        for toot in result:
            # Extract the toot_id from the current Toot object
            toot_id = toot.toot_id

            # Check if the toot_id is present in toots_dict
            self.assertTrue(any(toot_id == t.toot_id for t in toots_dict))

            if not any(toot_id == t.toot_id for t in toots_dict):
                true_bool = False

        for toot in toots_dict:
            # Extract the toot_id from the current Toot object
            toot_id = toot.toot_id

        # Check if the toot_id is present in result (at least once)
            self.assertTrue(any(toot_id == t.toot_id for t in result))

            if not any(toot_id == t.toot_id for t in result):
                true_bool = False

        # eine Nachricht basierend auf einer BOOL Variable!
        self.assertTrue(true_bool, "Your Loading-Function does not work correctly, check again if you load all Toots in the way they should!")

    def test_GetTextContent(self):
        text = 'Hello from Python, dog'
        text_content = MastodonOOPsolution.get_text_content(self.toot_true)
        self.assertEqual(
            text, text_content, "Your GetTextContent-Function does not work correctly, check again if only the text remains and no HTML is left!")

    def test_MediaTrigger(self):
        media = MastodonOOPsolution.MediaTrigger()

        self.assertTrue(issubclass(MastodonOOPsolution.MediaTrigger, MastodonOOPsolution.Trigger),
                        "MediaTrigger should be a subclass of Trigger")
        self.assertTrue(media.evaluate(self.toot_true), "Media-Trigger: Expected True but Output is False.")
        self.assertFalse(media.evaluate(self.toot_false), "Media-Trigger: Expected False but Output is True.")

    def test_ImageMediaTrigger(self):
        image = MastodonOOPsolution.ImageMediaTrigger()

        self.assertTrue(issubclass(MastodonOOPsolution.ImageMediaTrigger, MastodonOOPsolution.MediaTrigger),
                        "ImageMediaTrigger should be a subclass of MediaTrigger")
        self.assertTrue(image.evaluate(self.toot_true), "Image-Media-Trigger: Expected True but Output is False.")
        self.assertFalse(image.evaluate(self.toot_false), "Image-Media-Trigger: Expected False but Output is True.")

    def test_VideoMediaTrigger(self):
        video = MastodonOOPsolution.VideoMediaTrigger()

        self.assertTrue(issubclass(MastodonOOPsolution.VideoMediaTrigger, MastodonOOPsolution.MediaTrigger),
                        "VideoMediaTrigger should be a subclass of MediaTrigger")
        self.assertTrue(video.evaluate(self.toot_true), "Video-Media-Trigger: Expected True but Output is False.")
        self.assertFalse(video.evaluate(self.toot_false), "Video-Media-Trigger: Expected False but Output is True.")

    def test_GifMediaTrigger(self):
        gif = MastodonOOPsolution.GifMediaTrigger()

        self.assertTrue(issubclass(MastodonOOPsolution.GifMediaTrigger, MastodonOOPsolution.MediaTrigger),
                        "GifMediaTrigger should be a subclass of MediaTrigger")
        self.assertTrue(gif.evaluate(self.toot_true), "GIF-Media-Trigger: Expected True but Output is False.")
        self.assertFalse(gif.evaluate(self.toot_false), "GIF-Media-Trigger: Expected False but Output is True.")

    def test_AudioMediaTrigger(self):
        audio = MastodonOOPsolution.AudioMediaTrigger()

        self.assertTrue(issubclass(MastodonOOPsolution.AudioMediaTrigger, MastodonOOPsolution.MediaTrigger),
                        "AudioMediaTrigger should be a subclass of MediaTrigger")
        self.assertTrue(audio.evaluate(self.toot_true), "Audio-Media-Trigger: Expected True but Output is False.")
        self.assertFalse(audio.evaluate(self.toot_false), "Audio-Media-Trigger: Expected False but Output is True.")

    def test_LanguageTrigger(self):
        english = MastodonOOPsolution.LanguageTrigger("en")

        self.assertTrue(issubclass(MastodonOOPsolution.LanguageTrigger, MastodonOOPsolution.Trigger),
                        "LanguageTrigger should be a subclass of Trigger")
        self.assertTrue(english.evaluate(self.toot_true), "Language-Trigger: Expected True but Output is False.")
        self.assertFalse(english.evaluate(self.toot_false), "Language-Trigger: Expected False but Output is True.")

    def test_PollTrigger(self):
        poll_filter = MastodonOOPsolution.PollTrigger()

        self.assertTrue(issubclass(MastodonOOPsolution.PollTrigger, MastodonOOPsolution.Trigger),
                        "PollTrigger should be a subclass of Trigger")
        self.assertTrue(poll_filter.evaluate(self.toot_true), "Poll-Trigger: Expected True but Output is False.")
        self.assertFalse(poll_filter.evaluate(self.toot_false), "Poll-Trigger: Expected False but Output is True.")

    def test_MentionsTrigger(self):
        mentions = MastodonOOPsolution.MentionsTrigger()

        self.assertTrue(issubclass(MastodonOOPsolution.MentionsTrigger, MastodonOOPsolution.Trigger),
                        "MentionsTrigger should be a subclass of Trigger")
        self.assertTrue(mentions.evaluate(self.toot_true), "Mentions-Trigger: Expected True but Output is False.")
        self.assertFalse(mentions.evaluate(self.toot_false), "Mentions-Trigger: Expected False but Output is True.")

    def test_PhraseTrigger(self):
        phrase = MastodonOOPsolution.PhraseTrigger("dog")

        self.assertTrue(issubclass(MastodonOOPsolution.PhraseTrigger, MastodonOOPsolution.Trigger),
                        "PhraseTrigger should be a subclass of Trigger")
        self.assertTrue(phrase.evaluate(self.toot_true), "Phrase-Trigger: Expected True but Output is False.")
        self.assertFalse(phrase.evaluate(self.toot_false), "Phrase-Trigger: Expected False but Output is True.")

    def test_TimeTrigger(self):
        formatted_time = datetime.datetime.strptime('2023-07-22 09:37:34-05:00', "%Y-%m-%d %H:%M:%S%z")
        triggered_time = MastodonOOPsolution.TimeTrigger(self.clock)

        self.assertTrue(issubclass(MastodonOOPsolution.TimeTrigger, MastodonOOPsolution.Trigger),
                        "TimeTrigger should be a subclass of Trigger")
        self.assertEqual(triggered_time.ptime, formatted_time,
                         "Time-Trigger: The format of the trigger does not fit the wanted format YYYY-MM-DD hh:mm:ss+TZ'!")

    def test_BeforeTrigger(self):
        toot_true = self.toot_true
        toot_false = self.toot_false
        before = MastodonOOPsolution.BeforeTrigger(ptime=self.clock)

        self.assertTrue(issubclass(MastodonOOPsolution.BeforeTrigger, MastodonOOPsolution.TimeTrigger),
                        "BeforeTrigger should be a subclass of TimeTrigger")
        self.assertTrue(before.evaluate(toot_true), "Before-Trigger: Expected True but Output is False.")
        self.assertFalse(before.evaluate(toot_false), "Before-Trigger: Expected False but Output is True.")

    def test_AfterTrigger(self):
        test_clock = datetime.datetime.strptime(self.clock, "%Y-%m-%d %H:%M:%S%z")
        time = test_clock < self.toot_false.pubdate
        time2 = test_clock < self.toot_true.pubdate
        after = MastodonOOPsolution.AfterTrigger(self.clock)

        self.assertTrue(issubclass(MastodonOOPsolution.AfterTrigger, MastodonOOPsolution.TimeTrigger),
                        "AfterTrigger should be a subclass of TimeTrigger")
        self.assertEqual(after.evaluate(self.toot_false), time, "After-Trigger: Expected True but Output is False.")
        self.assertEqual(after.evaluate(self.toot_true), time2, "After-Trigger: Expected False but Output is True.")

    def test_NotTrigger(self):
        nottrigger = MastodonOOPsolution.NotTrigger(MastodonOOPsolution.MediaTrigger())

        self.assertTrue(issubclass(MastodonOOPsolution.NotTrigger, MastodonOOPsolution.Trigger),
                        "NotTrigger should be a subclass of Trigger")
        self.assertTrue(nottrigger.evaluate(self.toot_false), "Not-Trigger: Expected True but Output is False.")
        self.assertFalse(nottrigger.evaluate(self.toot_true), "Not-Trigger: Expected False but Output is True.")

    def test_AndTrigger(self):
        media = MastodonOOPsolution.MediaTrigger()
        gif = MastodonOOPsolution.GifMediaTrigger()
        andtrigger = MastodonOOPsolution.AndTrigger(
            trigger1=media,
            trigger2=gif
        )

        self.assertTrue(issubclass(MastodonOOPsolution.AndTrigger, MastodonOOPsolution.Trigger),
                        "AndTrigger should be a subclass of Trigger")
        self.assertTrue(andtrigger.evaluate(self.toot_true), "And-Trigger: Expected True but Output is False.")
        self.assertFalse(andtrigger.evaluate(self.toot_false), "And-Trigger: Expected False but Output is True.")

    def test_OrTrigger(self):
        poll = MastodonOOPsolution.PollTrigger()
        mentions = MastodonOOPsolution.MentionsTrigger()
        ortrigger = MastodonOOPsolution.OrTrigger(
            trigger1=poll,
            trigger2=mentions
        )

        self.assertTrue(issubclass(MastodonOOPsolution.OrTrigger, MastodonOOPsolution.Trigger),
                        "orTrigger should be a subclass of Trigger")
        self.assertTrue(ortrigger.evaluate(self.toot_true), "Or-Trigger: Expected True but Output is False.")
        self.assertFalse(ortrigger.evaluate(self.toot_false), "Or-Trigger: Expected False but Output is True.")

    def test_Filter(self):
        poll = MastodonOOPsolution.PollTrigger()
        mentions = MastodonOOPsolution.MentionsTrigger()
        media = MastodonOOPsolution.MediaTrigger()

        toot_list_true = [self.toot_true]
        toot_list_false = [self.toot_false]
        trigger_list = [poll, mentions, media]

        filter_true = MastodonOOPsolution.filter_toots(toots=toot_list_true, triggerlist=trigger_list)
        filter_false = MastodonOOPsolution.filter_toots(toots=toot_list_false, triggerlist=trigger_list)

        self.assertIsNotNone(filter_true, "Filter_Toots: Expected A but Output is None.")
        self.assertEqual(filter_true, toot_list_true, "Filter-Toots: Expected [A,B,C] but Output is not [A,B,C].")

        self.assertEqual(filter_false, [],
                         "Filter-Toots: Expected [] but Output is not [A,B,C], maybe check again if all of your trigger in the trigger-list get checked.")

    def test_Load_to_Workbook(self):
        temp_filename = 'test_objects.xlsx'
        workbook = openpyxl.Workbook()
        workbook.save(temp_filename)
        toot_list = [self.toot_true]

        MastodonOOPsolution.load_to_workbook(toot_list, temp_filename)

        saved_workbook = openpyxl.load_workbook(temp_filename)
        saved_worksheet = saved_workbook.active

        # Check if the written data matches our expectations
        self.assertEqual(saved_worksheet['A2'].value, toot_list[0].account[0]["username"],
                         "Load-To-Workbook: A2-Value differs from the value in the toot-list")
        self.assertEqual(saved_worksheet['B2'].value, toot_list[0].pubdate.replace(tzinfo=None),
                         "Load-To-Workbook: B2-Value differs from the value in the toot-list")
        self.assertEqual(saved_worksheet['C2'].value, toot_list[0].content,
                         "Load-To-Workbook: C2-Value differs from the value in the toot-list")

        saved_workbook.close()

        os.remove(temp_filename)


if __name__ == '__main__':
    unittest.main()